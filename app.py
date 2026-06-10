from calendar import monthrange
from datetime import date, datetime, timedelta
import os
from pathlib import Path

from flask import Flask, flash, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from db import (
    LOG_ACTIONS,
    STATUSES,
    count_cars,
    create_booking,
    fetch_booking,
    fetch_bookings,
    fetch_cars,
    fetch_logs,
    find_conflicts,
    init_db,
    log_activity,
    set_booking_status,
    update_booking,
)


app = Flask(__name__)
app.secret_key = "local-car-booking-secret"

IMPORT_FILE = Path("input/Резервирование автомобиля.xlsx")
EXPORT_FILE = Path("output/car_booking_report.xlsx")

STATUS_LABELS = {
    "confirmed": "Подтверждена",
    "active": "В поездке",
    "completed": "Завершена",
    "cancelled": "Отменена",
}

ACTION_LABELS = {
    "created": "Создана",
    "updated": "Изменена",
    "cancelled": "Отменена",
    "completed": "Завершена",
    "imported": "Импорт Excel",
    "exported": "Экспорт Excel",
}


def parse_date(value: str) -> date:
    """Преобразует дату из HTML-формы в объект date."""
    return datetime.strptime(value, "%Y-%m-%d").date()


def format_date_ru(value) -> str:
    """Показывает дату как ДД.ММ.ГГГГ."""
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d.%m.%Y")
        except ValueError:
            pass
    return text


def format_datetime_ru(value) -> str:
    """Показывает дату и время как ДД.ММ.ГГГГ ЧЧ:ММ."""
    if not value:
        return ""
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt == "%Y-%m-%d":
                return parsed.strftime("%d.%m.%Y")
            return parsed.strftime("%d.%m.%Y %H:%M")
        except ValueError:
            pass
    return text


def format_time(value) -> str:
    """Показывает время как ЧЧ:ММ."""
    if not value:
        return ""
    text = str(value).strip()
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).strftime("%H:%M")
        except ValueError:
            pass
    return text


def time_range_text(booking: dict) -> str:
    """Возвращает строку времени для карточек и календаря."""
    start_time = format_time(booking.get("start_time"))
    end_time = format_time(booking.get("end_time"))
    if start_time and end_time:
        return f"{start_time}-{end_time}"
    return ""


def days_count(start_date: str, end_date: str) -> int:
    """Считает количество календарных дней брони."""
    return (parse_date(end_date) - parse_date(start_date)).days + 1


def validate_time_pair(start_date: str, end_date: str, start_time: str, end_time: str) -> str | None:
    """Проверяет необязательные поля времени."""
    if bool(start_time) != bool(end_time):
        return "Укажите оба времени или оставьте оба поля пустыми."

    if not start_time and not end_time:
        return None

    try:
        datetime.strptime(start_time, "%H:%M")
        datetime.strptime(end_time, "%H:%M")
    except ValueError:
        return "Проверьте формат времени."

    if start_date == end_date and end_time < start_time:
        return "Время приезда не может быть раньше времени выезда."

    return None


def normalize_booking_form(form, is_edit: bool = False) -> tuple[dict, str | None]:
    """Собирает данные формы брони."""
    data = {
        "car_id": int(form.get("car_id", "0") or 0),
        "employee_name": form.get("employee_name", "").strip(),
        "start_date": form.get("start_date", "").strip(),
        "end_date": form.get("end_date", "").strip(),
        "start_time": form.get("start_time", "").strip(),
        "end_time": form.get("end_time", "").strip(),
        "route": form.get("route", "").strip(),
        "odometer_return": form.get("odometer_return", "").strip(),
        "status": form.get("status", "confirmed").strip(),
        "comment": form.get("comment", "").strip(),
        "actor_name": form.get("actor_name", "").strip(),
    }

    required = [
        data["car_id"],
        data["employee_name"],
        data["start_date"],
        data["end_date"],
        data["route"],
    ]
    if not all(required):
        return data, "Заполните все обязательные поля."

    try:
        start = parse_date(data["start_date"])
        end = parse_date(data["end_date"])
    except ValueError:
        return data, "Проверьте формат дат."

    if end < start:
        return data, "Дата приезда не может быть раньше даты выезда."

    time_error = validate_time_pair(data["start_date"], data["end_date"], data["start_time"], data["end_time"])
    if time_error:
        return data, time_error

    if data["status"] not in STATUSES:
        return data, "Выберите корректный статус."

    if not data["actor_name"]:
        data["actor_name"] = data["employee_name"]

    return data, None


def row_to_dict(row) -> dict:
    """Преобразует sqlite Row в словарь и добавляет расчетные поля."""
    item = dict(row)
    item["days_count"] = days_count(item["start_date"], item["end_date"])
    item["status_label"] = STATUS_LABELS.get(item["status"], item["status"])
    item["start_date_ru"] = format_date_ru(item["start_date"])
    item["end_date_ru"] = format_date_ru(item["end_date"])
    item["start_time_ru"] = format_time(item.get("start_time"))
    item["end_time_ru"] = format_time(item.get("end_time"))
    item["time_range"] = time_range_text(item)
    return item


def log_to_dict(row) -> dict:
    """Преобразует запись журнала для шаблона."""
    item = dict(row)
    item["created_at_ru"] = format_datetime_ru(item["created_at"])
    item["action_label"] = ACTION_LABELS.get(item["action"], item["action"])
    return item


def booking_summary(booking: dict) -> str:
    """Короткое описание брони для журнала."""
    time_part = f", время {booking['start_time']}-{booking['end_time']}" if booking.get("start_time") else ""
    return (
        f"{format_date_ru(booking['start_date'])} - {format_date_ru(booking['end_date'])}"
        f"{time_part}, маршрут: {booking.get('route', '')}"
    )


def build_update_details(old: dict, new: dict) -> str:
    """Формирует текст, что именно изменилось при редактировании."""
    labels = {
        "car_id": "автомобиль",
        "employee_name": "водитель",
        "start_date": "дата выезда",
        "end_date": "дата приезда",
        "start_time": "время выезда",
        "end_time": "время приезда",
        "route": "маршрут",
        "odometer_return": "одометр",
        "status": "статус",
        "comment": "комментарий",
    }
    changes = []
    for key, label in labels.items():
        old_value = str(old.get(key) or "")
        new_value = str(new.get(key) or "")
        if old_value != new_value:
            changes.append(f"{label}: '{old_value}' -> '{new_value}'")
    return "; ".join(changes) if changes else "Данные сохранены без изменений."


def month_shift(year: int, month: int, delta: int) -> tuple[int, int]:
    """Сдвигает год и месяц на delta месяцев."""
    month_index = year * 12 + month - 1 + delta
    return month_index // 12, month_index % 12 + 1


def detect_car_id_by_sheet(sheet_name: str) -> int | None:
    """Определяет одну из трех машин по названию листа старого Excel."""
    upper_name = sheet_name.upper()
    cars = fetch_cars()
    if "X585" in upper_name:
        return cars[0]["id"]
    if "X588" in upper_name or "Х588" in upper_name:
        return cars[1]["id"]
    if "TOYOTA" in upper_name or "C011" in upper_name:
        return cars[2]["id"]
    return None


def parse_excel_date(value) -> tuple[str | None, str | None]:
    """Пытается разобрать дату из старого Excel."""
    if value is None:
        return None, None
    if isinstance(value, datetime):
        return value.date().isoformat(), None
    if isinstance(value, date):
        return value.isoformat(), None

    text = str(value).strip()
    if not text:
        return None, None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat(), None
        except ValueError:
            pass
    return None, text


def is_header_or_month_row(values: list) -> bool:
    """Отсекает заголовки, месяцы и пустые строки старого Excel."""
    joined = " ".join(str(value).strip().lower() for value in values if value is not None)
    if not joined:
        return True
    skip_words = [
        "дата выезда",
        "дата приезда",
        "фио",
        "маршрут",
        "январ",
        "феврал",
        "март",
        "апрел",
        "май",
        "июн",
        "июл",
        "август",
        "сентябр",
        "октябр",
        "ноябр",
        "декабр",
    ]
    return any(word in joined for word in skip_words)


def import_old_excel() -> dict:
    """Импортирует старый Excel в SQLite без создания лишних машин."""
    result = {"found": 0, "imported": 0, "duplicates": 0, "errors": 0}
    if not IMPORT_FILE.exists():
        raise FileNotFoundError(f"Файл {IMPORT_FILE} не найден.")

    workbook = load_workbook(IMPORT_FILE, data_only=True)
    existing_keys = {
        (
            booking["car_id"],
            booking["start_date"],
            booking["end_date"],
            booking["employee_name"],
            booking["route"],
        )
        for booking in fetch_bookings()
    }

    for sheet in workbook.worksheets:
        car_id = detect_car_id_by_sheet(sheet.title)
        if car_id is None:
            continue

        for row in sheet.iter_rows(values_only=True):
            values = list(row[:6])
            if is_header_or_month_row(values):
                continue

            start_date, start_text = parse_excel_date(values[0] if len(values) > 0 else None)
            employee_name = str(values[2] or "").strip() if len(values) > 2 else ""
            route = str(values[3] or "").strip() if len(values) > 3 else ""
            if not start_date or not (employee_name or route):
                result["errors"] += 1
                continue

            result["found"] += 1
            end_date, end_text = parse_excel_date(values[1] if len(values) > 1 else None)
            comment_parts = []
            if start_text:
                comment_parts.append(f"Дата выезда из Excel: {start_text}")
            if end_text:
                comment_parts.append(f"Дата приезда из Excel: {end_text}")
            if not end_date:
                end_date = start_date

            odometer = str(values[4] or "").strip() if len(values) > 4 else ""
            note = str(values[5] or "").strip() if len(values) > 5 else ""
            if note:
                comment_parts.append(note)

            key = (car_id, start_date, end_date, employee_name, route)
            if key in existing_keys:
                result["duplicates"] += 1
                continue

            create_booking(
                {
                    "car_id": car_id,
                    "employee_name": employee_name,
                    "start_date": start_date,
                    "end_date": end_date,
                    "start_time": "",
                    "end_time": "",
                    "route": route,
                    "odometer_return": odometer,
                    "status": "completed",
                    "comment": "; ".join(comment_parts),
                }
            )
            existing_keys.add(key)
            result["imported"] += 1

    log_activity(
        "imported",
        actor_name="Система",
        details=(
            f"Импорт Excel: найдено {result['found']}, импортировано {result['imported']}, "
            f"дублей {result['duplicates']}, ошибок {result['errors']}"
        ),
    )
    return result


def add_month_separator(sheet, row_index: int, title: str, max_column: int) -> None:
    """Добавляет красную строку-разделитель месяца в Excel."""
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=max_column)
    cell = sheet.cell(row=row_index, column=1, value=title)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="B00020")
    cell.alignment = Alignment(horizontal="center")


def export_excel_report() -> Path:
    """Создает Excel-отчет с тремя листами по машинам."""
    EXPORT_FILE.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    headers = [
        "Дата выезда",
        "Время выезда",
        "Дата приезда",
        "Время приезда",
        "Ф.И.О. водителя",
        "Маршрут (адрес, наименование организации)",
        "Показание одометра при возвращении",
        "Примечания",
        "Статус",
    ]
    month_names = [
        "",
        "Январь",
        "Февраль",
        "Март",
        "Апрель",
        "Май",
        "Июнь",
        "Июль",
        "Август",
        "Сентябрь",
        "Октябрь",
        "Ноябрь",
        "Декабрь",
    ]
    border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    for car in fetch_cars():
        sheet = workbook.create_sheet(car["name"][:31])
        row_index = 1
        bookings = [row_to_dict(row) for row in fetch_bookings({"car_id": str(car["id"])})]
        bookings.sort(key=lambda item: (item["start_date"], item.get("start_time") or ""))
        current_month = None

        if not bookings:
            for column_index, header in enumerate(headers, start=1):
                cell = sheet.cell(row=row_index, column=column_index, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="F2F2F2")
                cell.border = border

        for booking in bookings:
            start = parse_date(booking["start_date"])
            month_key = (start.year, start.month)
            if current_month != month_key:
                add_month_separator(sheet, row_index, f"{month_names[start.month]} {start.year}", len(headers))
                row_index += 1
                for column_index, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=row_index, column=column_index, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="F2F2F2")
                    cell.border = border
                row_index += 1
                current_month = month_key

            values = [
                format_date_ru(booking["start_date"]),
                format_time(booking.get("start_time")),
                format_date_ru(booking["end_date"]),
                format_time(booking.get("end_time")),
                booking["employee_name"],
                booking["route"],
                booking["odometer_return"],
                booking["comment"],
                booking["status_label"],
            ]
            for column_index, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_index, column=column_index, value=value)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_index += 1

        widths = [14, 12, 14, 12, 24, 42, 24, 34, 16]
        for column_index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column_index)].width = width

    workbook.save(EXPORT_FILE)
    log_activity("exported", actor_name="Система", details=f"Выполнен экспорт Excel: {EXPORT_FILE.name}")
    return EXPORT_FILE


@app.context_processor
def inject_globals():
    """Передает общие значения и форматтеры во все шаблоны."""
    return {
        "status_labels": STATUS_LABELS,
        "statuses": STATUSES,
        "action_labels": ACTION_LABELS,
        "log_actions": LOG_ACTIONS,
        "format_date_ru": format_date_ru,
        "format_datetime_ru": format_datetime_ru,
        "format_time": format_time,
        "time_range_text": time_range_text,
    }


@app.route("/")
def index():
    """Главная страница со списком броней и фильтрами."""
    init_db()
    filters = {
        "car_id": request.args.get("car_id", "").strip(),
        "date_from": request.args.get("date_from", "").strip(),
        "date_to": request.args.get("date_to", "").strip(),
        "status": request.args.get("status", "").strip(),
        "search": request.args.get("search", "").strip(),
    }
    bookings = [row_to_dict(row) for row in fetch_bookings(filters)]
    today = date.today().isoformat()
    all_bookings = fetch_bookings()
    active_count = len([booking for booking in all_bookings if booking["status"] in ("confirmed", "active")])
    busy_today = {
        booking["car_id"]
        for booking in all_bookings
        if booking["status"] in ("confirmed", "active")
        and booking["start_date"] <= today <= booking["end_date"]
    }
    upcoming = [
        row_to_dict(booking)
        for booking in all_bookings
        if booking["status"] in ("confirmed", "active") and booking["end_date"] >= today
    ]
    nearest = sorted(upcoming, key=lambda item: (item["start_date"], item.get("start_time") or ""))[0] if upcoming else None

    stats = {
        "total": len(all_bookings),
        "active": active_count,
        "free_today": max(count_cars() - len(busy_today), 0),
        "nearest": nearest,
    }
    return render_template("index.html", bookings=bookings, cars=fetch_cars(), filters=filters, stats=stats)


@app.route("/booking/new", methods=["GET", "POST"])
def new_booking():
    """Создание новой брони."""
    conflicts = []
    form_data = {}
    if request.method == "POST":
        data, error = normalize_booking_form(request.form)
        form_data = data
        if error:
            flash(error, "error")
        else:
            conflicts = [
                row_to_dict(row)
                for row in find_conflicts(
                    data["car_id"],
                    data["start_date"],
                    data["end_date"],
                    data["start_time"],
                    data["end_time"],
                )
            ]
            if conflicts:
                flash("Машина занята на пересекающиеся даты или время. Бронь не создана.", "error")
            else:
                data["status"] = "confirmed"
                booking_id = create_booking(data)
                booking = row_to_dict(fetch_booking(booking_id))
                log_activity(
                    "created",
                    booking_id=booking_id,
                    car_name=booking["car_name"],
                    actor_name=data["actor_name"],
                    employee_name=booking["employee_name"],
                    details=f"Создана бронь: {booking_summary(booking)}",
                )
                flash("Бронь подтверждена.", "success")
                return redirect(url_for("edit_booking", booking_id=booking_id))

    return render_template(
        "booking_form.html",
        title="Новая бронь",
        action_url=url_for("new_booking"),
        cars=fetch_cars(),
        booking=form_data,
        conflicts=conflicts,
        is_edit=False,
    )


@app.route("/booking/<int:booking_id>/edit", methods=["GET", "POST"])
def edit_booking(booking_id: int):
    """Редактирование брони с повторной проверкой конфликтов."""
    booking = fetch_booking(booking_id)
    if booking is None:
        flash("Бронь не найдена.", "error")
        return redirect(url_for("index"))

    conflicts = []
    form_data = dict(booking)
    if request.method == "POST":
        old_booking = dict(booking)
        data, error = normalize_booking_form(request.form, is_edit=True)
        form_data = data
        if error:
            flash(error, "error")
        else:
            if data["status"] in ("confirmed", "active"):
                conflicts = [
                    row_to_dict(row)
                    for row in find_conflicts(
                        data["car_id"],
                        data["start_date"],
                        data["end_date"],
                        data["start_time"],
                        data["end_time"],
                        booking_id,
                    )
                ]
            if conflicts:
                flash("Машина занята на пересекающиеся даты или время. Изменения не сохранены.", "error")
            else:
                update_booking(booking_id, data)
                updated = row_to_dict(fetch_booking(booking_id))
                log_activity(
                    "updated",
                    booking_id=booking_id,
                    car_name=updated["car_name"],
                    actor_name=data["actor_name"],
                    employee_name=updated["employee_name"],
                    details=build_update_details(old_booking, data),
                )
                flash("Бронь обновлена.", "success")
                return redirect(url_for("index"))

    return render_template(
        "booking_form.html",
        title=f"Редактирование брони #{booking_id}",
        action_url=url_for("edit_booking", booking_id=booking_id),
        cars=fetch_cars(),
        booking=form_data,
        conflicts=conflicts,
        is_edit=True,
    )


@app.route("/booking/<int:booking_id>/cancel", methods=["GET", "POST"])
def cancel_booking(booking_id: int):
    """Отмена брони через форму подтверждения."""
    booking = fetch_booking(booking_id)
    if booking is None:
        flash("Бронь не найдена.", "error")
        return redirect(url_for("index"))

    booking_dict = row_to_dict(booking)
    if request.method == "POST":
        actor_name = request.form.get("actor_name", "").strip() or booking["employee_name"]
        reason = request.form.get("reason", "").strip()
        if not reason:
            flash("Укажите причину отмены.", "error")
        else:
            set_booking_status(booking_id, "cancelled", extra_comment=f"Причина отмены: {reason}")
            log_activity(
                "cancelled",
                booking_id=booking_id,
                car_name=booking["car_name"],
                actor_name=actor_name,
                employee_name=booking["employee_name"],
                details=f"Бронь отменена. Причина: {reason}",
            )
            flash("Бронь отменена.", "success")
            return redirect(url_for("index"))

    return render_template("cancel_booking.html", booking=booking_dict)


@app.route("/booking/<int:booking_id>/complete", methods=["GET", "POST"])
def complete_booking(booking_id: int):
    """Завершение поездки с вводом одометра и автора действия."""
    booking = fetch_booking(booking_id)
    if booking is None:
        flash("Бронь не найдена.", "error")
        return redirect(url_for("index"))

    booking_dict = row_to_dict(booking)
    if request.method == "POST":
        actor_name = request.form.get("actor_name", "").strip() or booking["employee_name"]
        odometer = request.form.get("odometer_return", "").strip()
        completion_comment = request.form.get("completion_comment", "").strip()
        if not odometer:
            flash("Введите показание одометра.", "error")
        else:
            extra_comment = f"Комментарий завершения: {completion_comment}" if completion_comment else None
            set_booking_status(booking_id, "completed", odometer, extra_comment)
            log_activity(
                "completed",
                booking_id=booking_id,
                car_name=booking["car_name"],
                actor_name=actor_name,
                employee_name=booking["employee_name"],
                details=f"Поездка завершена. Одометр: {odometer}. {completion_comment}",
            )
            flash("Поездка завершена.", "success")
            return redirect(url_for("index"))

    return render_template("complete_booking.html", booking=booking_dict)


@app.route("/calendar")
def calendar_view():
    """Календарь на месяц: автомобили строками, дни колонками."""
    init_db()
    today = date.today()
    year = int(request.args.get("year", today.year))
    month = int(request.args.get("month", today.month))
    days_in_month = monthrange(year, month)[1]
    days = [date(year, month, day) for day in range(1, days_in_month + 1)]
    first_day = days[0].isoformat()
    last_day = days[-1].isoformat()
    bookings = [row_to_dict(row) for row in fetch_bookings({"date_from": first_day, "date_to": last_day})]
    cars = fetch_cars()
    grid = {car["id"]: {day.isoformat(): [] for day in days} for car in cars}

    for booking in bookings:
        start = max(parse_date(booking["start_date"]), days[0])
        end = min(parse_date(booking["end_date"]), days[-1])
        current = start
        while current <= end:
            grid[booking["car_id"]][current.isoformat()].append(booking)
            current += timedelta(days=1)

    prev_year, prev_month = month_shift(year, month, -1)
    next_year, next_month = month_shift(year, month, 1)
    return render_template(
        "calendar.html",
        cars=cars,
        days=days,
        grid=grid,
        year=year,
        month=month,
        prev_year=prev_year,
        prev_month=prev_month,
        next_year=next_year,
        next_month=next_month,
        today=today,
    )


@app.route("/logs")
def logs_page():
    """Журнал действий."""
    filters = {
        "action": request.args.get("action", "").strip(),
        "car_name": request.args.get("car_name", "").strip(),
        "date_from": request.args.get("date_from", "").strip(),
        "date_to": request.args.get("date_to", "").strip(),
        "search": request.args.get("search", "").strip(),
    }
    logs = [log_to_dict(row) for row in fetch_logs(filters)]
    return render_template("logs.html", logs=logs, cars=fetch_cars(), filters=filters)


@app.route("/import", methods=["GET", "POST"])
def import_page():
    """Страница импорта старого Excel."""
    result = None
    if request.method == "POST":
        try:
            result = import_old_excel()
            flash("Импорт завершен.", "success")
        except FileNotFoundError as error:
            flash(str(error), "error")
    return render_template("import_excel.html", import_file=IMPORT_FILE, result=result, cars_count=count_cars())


@app.route("/export")
def export_page():
    """Создает Excel-отчет и показывает ссылку на скачивание."""
    export_excel_report()
    return render_template("export_excel.html", export_file=EXPORT_FILE)


@app.route("/download/export")
def download_export():
    """Скачивание последнего Excel-отчета."""
    if not EXPORT_FILE.exists():
        export_excel_report()
    return send_file(EXPORT_FILE, as_attachment=True)


if __name__ == "__main__":
    init_db()
    port_from_env = os.environ.get("PORT")
    if port_from_env:
        app.run(host="0.0.0.0", port=int(port_from_env), debug=False)
    else:
        app.run(host="127.0.0.1", port=5000, debug=True)
